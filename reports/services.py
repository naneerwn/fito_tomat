from __future__ import annotations

import json
from pathlib import Path
import io
from datetime import datetime
from typing import Any, Dict, List, TypedDict, cast

from django.conf import settings
from django.db.models import Avg, Count, F
from django.db.models.functions import TruncDate
from django.utils import timezone

from diagnostics.models import Diagnosis, Image
from operations.models import Recommendation, Task
from common.typing import DateTimeLike
from infrastructure.models import Section, Greenhouse
from users.models import User
from .models import Report


class DiseaseDistributionEntry(TypedDict):
    disease__name: str
    total: int


class DiagnosticsPayload(TypedDict):
    total: int
    avg_confidence: float | None
    distribution: List[DiseaseDistributionEntry]


class RecommendationsPayload(TypedDict):
    total: int


class TasksPayload(TypedDict):
    total: int
    completed_on_time: int
    overdue: int


class PeriodPayload(TypedDict):
    start: str
    end: str


class ReportPayload(TypedDict, total=False):
    period: PeriodPayload
    diagnostics: DiagnosticsPayload
    recommendations: RecommendationsPayload
    tasks: TasksPayload
    timeseries: List[Dict[str, Any]]
    greenhouse_stats: List[Dict[str, Any]]
    operator_stats: List[Dict[str, Any]]
    economics: Dict[str, float]


def generate_report_payload(
    period_start: DateTimeLike,
    period_end: DateTimeLike,
) -> ReportPayload:
    """
    Aggregate KPI metrics for the requested period.
    """
    current_tz = timezone.get_current_timezone()
    start_dt = cast(datetime, period_start)
    end_dt = cast(datetime, period_end)
    start_dt = timezone.make_aware(start_dt, current_tz) if timezone.is_naive(start_dt) else start_dt
    end_dt = timezone.make_aware(end_dt, current_tz) if timezone.is_naive(end_dt) else end_dt

    diagnoses_qs = Diagnosis.objects.filter(timestamp__range=(start_dt, end_dt))
    tasks_qs = Task.objects.filter(created_at__range=(start_dt, end_dt))
    recommendations_qs = Recommendation.objects.filter(created_at__range=(start_dt, end_dt))

    disease_distribution = list(
        diagnoses_qs.values('disease__name')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    avg_confidence = diagnoses_qs.aggregate(avg=Avg('confidence'))['avg']

    completed_on_time = tasks_qs.filter(completed_at__isnull=False, completed_at__lte=F('deadline')).count()
    overdue = tasks_qs.filter(deadline__lt=timezone.now(), completed_at__isnull=True).count()

    # Временные ряды по датам
    timeseries = list(
        diagnoses_qs.annotate(dt=TruncDate('timestamp'))
        .values('dt')
        .annotate(total=Count('id'))
        .order_by('dt')
    )

    # Статистика по теплицам/секциям
    greenhouse_stats = list(
        diagnoses_qs.values('image__section__greenhouse__name', 'image__section__name')
        .annotate(total=Count('id'))
        .order_by('-total')
    )

    # Эффективность операторов (по задачам)
    operator_stats = list(
        tasks_qs.values('operator__full_name', 'operator__username')
        .annotate(total=Count('id'))
        .order_by('-total')
    )

    # Простейшие экономические оценки (плейсхолдеры, можно настроить)
    prevented_loss = diagnoses_qs.count() * 1.5  # условные тонны/кг
    saved_hours = tasks_qs.count() * 0.5  # условные часы экономии

    payload: ReportPayload = {
        'period': {
            'start': start_dt.isoformat(),
            'end': end_dt.isoformat(),
        },
        'diagnostics': {
            'total': diagnoses_qs.count(),
            'avg_confidence': round(avg_confidence, 4) if avg_confidence is not None else None,
            'distribution': disease_distribution,
        },
        'recommendations': {
            'total': recommendations_qs.count(),
        },
        'tasks': {
            'total': tasks_qs.count(),
            'completed_on_time': completed_on_time,
            'overdue': overdue,
        },
        'timeseries': [
            {'date': str(entry['dt']), 'total': entry['total']}
            for entry in timeseries
        ],
        'greenhouse_stats': [
            {
                'greenhouse': entry['image__section__greenhouse__name'] or '—',
                'section': entry['image__section__name'] or '—',
                'total': entry['total'],
            }
            for entry in greenhouse_stats
        ],
        'operator_stats': [
            {
                'operator': entry['operator__full_name'] or entry['operator__username'] or '—',
                'total': entry['total'],
            }
            for entry in operator_stats
        ],
        'economics': {
            'prevented_loss': round(prevented_loss, 2),
            'saved_hours': round(saved_hours, 2),
        },
    }
    return payload


def build_report_payload_by_type(
    period_start: DateTimeLike,
    period_end: DateTimeLike,
    report_type: str,
) -> ReportPayload:
    """
    Возвращает данные отчёта в зависимости от выбранного типа:
    - diagnostics_summary: только блоки диагностики (дистрибуция, временные ряды, теплицы)
    - tasks_summary: только блоки задач/операторов (и счётчики рекомендаций/задач)
    - full_report: все блоки
    """
    base = generate_report_payload(period_start, period_end)

    if report_type == 'diagnostics_summary':
        return {
            'period': base['period'],
            'diagnostics': base['diagnostics'],
            'timeseries': base.get('timeseries', []),
            'greenhouse_stats': base.get('greenhouse_stats', []),
        }

    if report_type == 'tasks_summary':
        return {
            'period': base['period'],
            'recommendations': base['recommendations'],
            'tasks': base['tasks'],
            'operator_stats': base.get('operator_stats', []),
        }

    # full_report или любой другой тип по умолчанию — возвращаем всё
    return base


def try_import_excel():
    from openpyxl import Workbook  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore

    return Workbook, get_column_letter, Alignment, Border, Font, PatternFill, Side


def try_import_pdf():
    from reportlab.lib.pagesizes import A4, landscape  # type: ignore
    from reportlab.lib import colors  # type: ignore
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer  # type: ignore
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle  # type: ignore
    from reportlab.pdfbase import pdfmetrics  # type: ignore
    from reportlab.pdfbase.ttfonts import TTFont  # type: ignore

    return (
        A4,
        landscape,
        colors,
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
        getSampleStyleSheet,
        ParagraphStyle,
        pdfmetrics,
        TTFont,
    )


def resolve_pdf_font(pdfmetrics, TTFont) -> str:
    font_candidates = [
        Path("C:/Windows/Fonts/Arial/ArialUnicodeMS.ttf"),
        Path("C:/Windows/Fonts/Arial/ARIALUNI.TTF"),
        Path("C:/Windows/Fonts/Arial/Arial.ttf"),
        Path("C:/Windows/Fonts/arialuni.ttf"),
        Path("C:/Windows/Fonts/ARIALUNI.TTF"),
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"),
    ]
    for candidate in font_candidates:
        if candidate.exists():
            try:
                pdfmetrics.registerFont(TTFont("CustomBase", str(candidate)))
                return "CustomBase"
            except Exception:
                continue
    return "Helvetica"


def build_excel_report(
    report: Report,
    details: Dict[str, Any],
    summary: Dict[str, Any],
    reporter: str,
    reporter_role: str,
):
    Workbook, get_column_letter, Alignment, Border, Font, PatternFill, Side = try_import_excel()

    include_diag = report.report_type in ('diagnostics_summary', 'full_report')
    include_tasks = report.report_type in ('tasks_summary', 'full_report')
    include_analytics = report.report_type == 'full_report'

    wb = Workbook()
    thin = Side(border_style="thin", color="000000")
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    report_title = {
        'diagnostics_summary': 'Сводка по диагностике',
        'tasks_summary': 'Сводка по задачам',
        'full_report': 'Полный отчёт',
    }.get(report.report_type, report.report_type)

    def style_header_row(ws, row_idx: int, col_count: int) -> None:
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    def apply_table_borders(ws, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                ws.cell(row=r, column=c).border = Border(top=thin, bottom=thin, left=thin, right=thin)

    def add_sheet_header(ws, col_count: int) -> int:
        ws.append([report_title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="left", vertical="center")

        ws.append([f"Сформирован: {report.generated_at:%Y-%m-%d %H:%M}, {reporter} ({reporter_role})"])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
        ws['A2'].alignment = Alignment(horizontal="left")

        ws.append([f"Период: {report.period_start:%Y-%m-%d %H:%M} — {report.period_end:%Y-%m-%d %H:%M}"])
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=col_count)
        ws['A3'].alignment = Alignment(horizontal="left")

        ws.append([])
        return 4

    current_sheet = None

    if include_diag:
        ws_diag = wb.active
        ws_diag.title = 'Диагностики'
        header_start_row = add_sheet_header(ws_diag, 9)
        ws_diag.append([
            'ID Диагноза', 'Дата/время', 'Теплица', 'Секция',
            'Изображение (путь)', 'Заболевание', 'Точность (%)',
            'Статус', 'Агроном'
        ])
        style_header_row(ws_diag, header_start_row, 9)
        for row in details['diagnoses']:
            ws_diag.append([
                row['id'],
                row['timestamp'].strftime('%Y-%m-%d %H:%M:%S') if row['timestamp'] else '',
                row['greenhouse'],
                row['section'],
                row['image_str'],
                row['disease'],
                row['confidence'],
                row['status'],
                row['verified_by'],
            ])
        apply_table_borders(ws_diag, header_start_row, 1, ws_diag.max_row, 9)
        for col in range(1, 10):
            ws_diag.column_dimensions[get_column_letter(col)].width = 18
        current_sheet = ws_diag

    if include_tasks:
        ws_rt = wb.create_sheet('Рекомендации и задачи') if current_sheet else wb.active
        header_start_row = add_sheet_header(ws_rt, 8)
        ws_rt.append([
            'ID Рекомендации', 'План лечения',
            'ID Задачи', 'Описание задачи', 'Исполнитель', 'Статус задачи',
            'Срок выполнения', 'Факт выполнения'
        ])
        style_header_row(ws_rt, header_start_row, 8)
        for row in details['rec_tasks']:
            ws_rt.append([
                row['rec_id'],
                row['plan'],
                row['task_id'],
                row['task_desc'],
                row['operator'],
                row['task_status'],
                row['deadline'].strftime('%Y-%m-%d %H:%M:%S') if row['deadline'] else '',
                row['completed_at'].strftime('%Y-%m-%d %H:%M:%S') if row['completed_at'] else '',
            ])
        apply_table_borders(ws_rt, header_start_row, 1, ws_rt.max_row, 8)
        for col in range(1, 9):
            ws_rt.column_dimensions[get_column_letter(col)].width = 20
        if not current_sheet:
            current_sheet = ws_rt

    if include_analytics:
        ws_an = wb.create_sheet('Аналитика')
        header_start_row = add_sheet_header(ws_an, 2)
        ws_an.append(['Метрика', 'Значение'])
        style_header_row(ws_an, header_start_row, 2)
        ws_an.append(['Всего диагнозов', summary['diagnostics']['total']])
        ws_an.append(['Средняя точность', summary['diagnostics']['avg_confidence'] or 0])
        ws_an.append(['Всего рекомендаций', summary['recommendations']['total']])
        ws_an.append(['Всего задач', summary['tasks']['total']])
        ws_an.append(['Выполнено в срок', summary['tasks']['completed_on_time']])
        ws_an.append(['Просрочено', summary['tasks']['overdue']])
        economics = summary.get('economics', {})
        ws_an.append(['Предотвращенные потери (условн.)', economics.get('prevented_loss', 0)])
        ws_an.append(['Экономия трудозатрат (час)', economics.get('saved_hours', 0)])
        apply_table_borders(ws_an, header_start_row, 1, ws_an.max_row, 2)
        ws_an.column_dimensions[get_column_letter(1)].width = 35
        ws_an.column_dimensions[get_column_letter(2)].width = 25

    return wb


def build_pdf_report(
    report: Report,
    details: Dict[str, Any],
    summary: Dict[str, Any],
    reporter: str,
    reporter_role: str,
) -> bytes:
    (
        A4,
        landscape,
        colors,
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
        getSampleStyleSheet,
        ParagraphStyle,
        pdfmetrics,
        TTFont,
    ) = try_import_pdf()

    include_diag = report.report_type in ('diagnostics_summary', 'full_report')
    include_tasks = report.report_type in ('tasks_summary', 'full_report')
    include_analytics = report.report_type == 'full_report'

    font_name = resolve_pdf_font(pdfmetrics, TTFont)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        title=f"Report #{report.id}",
        leftMargin=24,
        rightMargin=24,
        topMargin=36,
        bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    for style_name in ['Normal', 'Title', 'Heading1', 'Heading2', 'Heading3', 'BodyText']:
        if style_name in styles:
            styles[style_name].fontName = font_name
    styles.add(ParagraphStyle(name="Small", parent=styles['Normal'], fontSize=9))
    styles.add(ParagraphStyle(name="Header", parent=styles['Heading2'], fontSize=14, spaceAfter=8))
    styles.add(ParagraphStyle(name="TableCell", parent=styles['Normal'], fontSize=8, leading=10, wordWrap='CJK'))

    def wrap_cell(val):
        return Paragraph(str(val), styles['TableCell'])

    elems = []
    report_title = {
        'diagnostics_summary': 'Сводка по диагностике',
        'tasks_summary': 'Сводка по задачам',
        'full_report': 'Полный отчёт',
    }.get(report.report_type, report.report_type)
    elems.append(Paragraph(report_title, styles['Header']))
    elems.append(Paragraph(f"Сформирован: {report.generated_at:%Y-%m-%d %H:%M} • {reporter} ({reporter_role})", styles['Small']))
    elems.append(Paragraph(f"Период: {report.period_start:%Y-%m-%d %H:%M} — {report.period_end:%Y-%m-%d %H:%M}", styles['Small']))
    elems.append(Spacer(1, 12))

    def make_table(data, col_widths=None):
        tbl = Table(data, repeatRows=1, colWidths=col_widths)
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
        ]))
        return tbl

    if include_diag:
        diag_data = [["ID", "Дата/время", "Теплица", "Секция", "Изображение", "Заболевание", "Точность (%)", "Статус", "Агроном"]]
        for row in details['diagnoses']:
            diag_data.append([
                wrap_cell(row['id']),
                wrap_cell(row['timestamp'].strftime('%Y-%m-%d %H:%M:%S') if row['timestamp'] else ''),
                wrap_cell(row['greenhouse']),
                wrap_cell(row['section']),
                wrap_cell(row['image_str']),
                wrap_cell(row['disease']),
                wrap_cell(row['confidence']),
                wrap_cell(row['status']),
                wrap_cell(row['verified_by']),
            ])
        elems.append(Paragraph("Результаты диагностики", styles['Heading2']))
        elems.append(make_table(diag_data, col_widths=[45, 65, 65, 65, 150, 100, 60, 70, 70]))
        elems.append(Spacer(1, 12))

    if include_tasks:
        rt_data = [["ID Рекомендации", "План лечения", "ID Задачи", "Описание задачи", "Исполнитель", "Статус задачи", "Срок", "Факт"]]
        for row in details['rec_tasks']:
            rt_data.append([
                wrap_cell(row['rec_id']),
                wrap_cell(row['plan']),
                wrap_cell(row['task_id']),
                wrap_cell(row['task_desc']),
                wrap_cell(row['operator']),
                wrap_cell(row['task_status']),
                wrap_cell(row['deadline'].strftime('%Y-%m-%d %H:%M:%S') if row['deadline'] else ''),
                wrap_cell(row['completed_at'].strftime('%Y-%m-%d %H:%M:%S') if row['completed_at'] else ''),
            ])
        elems.append(Paragraph("Рекомендации и задачи", styles['Heading2']))
        elems.append(make_table(rt_data, col_widths=[80, 120, 60, 140, 80, 70, 70, 70]))
        elems.append(Spacer(1, 12))

    if include_analytics:
        anal_data = [
            ["Всего диагнозов", summary['diagnostics']['total']],
            ["Средняя точность", summary['diagnostics']['avg_confidence'] or 0],
            ["Всего рекомендаций", summary['recommendations']['total']],
            ["Всего задач", summary['tasks']['total']],
            ["Выполнено в срок", summary['tasks']['completed_on_time']],
            ["Просрочено", summary['tasks']['overdue']],
            ["Предотвращенные потери (условн.)", summary.get('economics', {}).get('prevented_loss', 0)],
            ["Экономия трудозатрат (час)", summary.get('economics', {}).get('saved_hours', 0)],
        ]
        anal_rows = [[wrap_cell(a), wrap_cell(b)] for a, b in anal_data]
        elems.append(Paragraph("Аналитический раздел", styles['Heading2']))
        elems.append(make_table(anal_rows, col_widths=[210, 140]))

    doc.build(elems)
    pdf = buf.getvalue()
    buf.close()
    return pdf


def fetch_detailed_data(
    period_start: DateTimeLike,
    period_end: DateTimeLike,
) -> Dict[str, Any]:
    """Возвращает детализированные данные для табличных частей отчёта."""
    current_tz = timezone.get_current_timezone()
    start_dt = cast(datetime, period_start)
    end_dt = cast(datetime, period_end)
    start_dt = timezone.make_aware(start_dt, current_tz) if timezone.is_naive(start_dt) else start_dt
    end_dt = timezone.make_aware(end_dt, current_tz) if timezone.is_naive(end_dt) else end_dt

    diagnoses_qs = (
        Diagnosis.objects.select_related(
            'image__section__greenhouse',
            'disease',
            'ml_disease',
            'verified_by',
        )
        .filter(timestamp__range=(start_dt, end_dt))
        .order_by('-timestamp')
    )

    def status_label(d: Diagnosis) -> str:
        if d.is_verified:
            if d.ml_disease and d.ml_disease_id != d.disease_id:
                return 'Скорректирован'
            return 'Подтвержден'
        return 'Ожидает проверки'

    diagnoses_rows = []
    for d in diagnoses_qs:
        section = getattr(d.image, 'section', None)
        greenhouse = getattr(section, 'greenhouse', None) if section else None
        diagnoses_rows.append({
            'id': d.id,
            'timestamp': d.timestamp,
            'greenhouse': greenhouse.name if greenhouse else '',
            'section': section.name if section else '',
            'image_path': getattr(d.image, 'file_path', ''),
            'image_str': str(getattr(d.image, 'file_path', '')) if getattr(d.image, 'file_path', None) else '',
            'disease': d.disease.name if d.disease else '',
            'confidence': round((d.confidence or 0) * 100, 2),
            'status': status_label(d),
            'verified_by': d.verified_by.full_name if d.verified_by and getattr(d.verified_by, 'full_name', '') else (d.verified_by.username if d.verified_by else ''),
            'note': '',
        })

    recommendations_qs = (
        Recommendation.objects.select_related(
            'diagnosis__image__section__greenhouse',
            'agronomist',
        )
        .filter(created_at__range=(start_dt, end_dt))
        .order_by('-created_at')
    )
    tasks_qs = (
        Task.objects.select_related('recommendation__diagnosis', 'operator')
        .filter(created_at__range=(start_dt, end_dt))
        .order_by('-created_at')
    )
    # Индексируем задачи по recommendation_id
    tasks_by_rec: Dict[int, List[Task]] = {}
    for t in tasks_qs:
        tasks_by_rec.setdefault(t.recommendation_id, []).append(t)

    rec_task_rows = []
    for rec in recommendations_qs:
        rec_tasks = tasks_by_rec.get(rec.id, [None])
        for t in rec_tasks:
            rec_task_rows.append({
                'rec_id': rec.id,
                'plan': rec.treatment_plan_text,
                'task_id': t.id if t else '',
                'task_desc': t.description if t else '',
                'operator': (t.operator.full_name if t and t.operator and getattr(t.operator, 'full_name', '') else (t.operator.username if t and t.operator else '')) if t else '',
                'task_status': t.status if t else '',
                'deadline': t.deadline if t else '',
                'completed_at': t.completed_at if t else '',
            })

    return {
        'diagnoses': diagnoses_rows,
        'rec_tasks': rec_task_rows,
    }


def persist_report_file(report_id: int, payload: Dict[str, Any]) -> Path:
    file_path = settings.REPORTS_DIR / f'report_{report_id}.json'
    file_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
    return file_path

